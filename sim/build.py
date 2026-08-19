"""
Build the BarMate dataset.

    python -m sim.build [--weather data/external/weather.csv]

Writes two directories:

  data/public/         everything the agent is allowed to see
  data/ground_truth/   held-out truth, used only for evaluation

Nothing in ground_truth is ever loaded into Supabase or Pinecone. If the agent
can read it, the evaluation is worthless.
"""
import argparse
import csv
import datetime
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from . import config as C
from . import incidents as INC
from . import master_data as M
from .engine import Simulation
from .reports import ReportWriter
from .validate import book_stock, truth_at

ROOT = Path(__file__).resolve().parent.parent
PUBLIC = ROOT / "data" / "public"
GT = ROOT / "data" / "ground_truth"
BROADCAST_XLSX = ROOT / "data" / "source" / "israeli_sports_broadcasts_only_week.xlsx"


def write_csv(path, rows, fieldnames=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("")
        return 0
    fieldnames = fieldnames or list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def load_broadcasts():
    """
    REAL data. Sourced from Sport5 and livegames.co.il, with the originating
    URL preserved per row. Not simulated and not extended.
    """
    wb = openpyxl.load_workbook(BROADCAST_XLSX)
    ws = wb["Broadcasts"]
    hdr = list(next(ws.iter_rows(values_only=True)))
    rows, by_date = [], defaultdict(list)
    for raw in list(ws.iter_rows(values_only=True))[1:]:
        r = dict(zip(hdr, raw))
        r["broadcast_date"] = str(r["broadcast_date"])[:10]
        rows.append(r)
        by_date[datetime.date.fromisoformat(r["broadcast_date"])].append(r)
    return rows, by_date


def load_optional_csv(path, key="date"):
    if not path or not Path(path).exists():
        return {}, []
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    out = {}
    for r in rows:
        for k, v in list(r.items()):
            if k != key:
                try:
                    r[k] = float(v)
                except (TypeError, ValueError):
                    pass
        out[datetime.date.fromisoformat(r[key])] = r
    return out, rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--weather", default=str(ROOT / "data" / "external" / "weather.csv"))
    ap.add_argument("--holidays", default=str(ROOT / "data" / "external" / "holidays.csv"))
    args = ap.parse_args()

    broadcast_rows, broadcasts = load_broadcasts()
    weather, weather_rows = load_optional_csv(args.weather)
    holiday_map, holiday_rows = load_optional_csv(args.holidays)
    holidays = {d: r.get("title", "") for d, r in holiday_map.items()}

    print(f"external context: {len(broadcast_rows)} broadcasts (real), "
          f"{len(weather_rows)} weather days, {len(holiday_rows)} holiday days")
    if not weather_rows:
        print("  note: weather absent, weather multipliers disabled. "
              "Run scripts/fetch_external.py to populate.")

    sim = Simulation(broadcasts=broadcasts, weather=weather, holidays=holidays).run()
    writer = ReportWriter(sim)
    shift_reports, shift_truth = writer.build()
    whatsapp = writer.whatsapp()

    # ------------------------------------------------------------ public
    counts = [{k: v for k, v in c.items()} for c in sim.counts]
    orders = [{k: v for k, v in o.items()} for o in sim.orders]

    written = {}
    written["products.csv"] = write_csv(PUBLIC / "products.csv", sim.products)
    written["suppliers.csv"] = write_csv(PUBLIC / "suppliers.csv", M.SUPPLIERS)
    written["cocktails.csv"] = write_csv(PUBLIC / "cocktails.csv", [
        {"cocktail_id": c["cocktail_id"], "name": c["name"],
         "name_he": c["name_he"], "price": c["price"]} for c in M.COCKTAILS])
    written["cocktail_recipes.csv"] = write_csv(PUBLIC / "cocktail_recipes.csv", sim.recipes)
    written["staff.csv"] = write_csv(PUBLIC / "staff.csv", M.STAFF)
    written["staff_schedule.csv"] = write_csv(PUBLIC / "staff_schedule.csv", sim.schedule)
    written["sales.csv"] = write_csv(PUBLIC / "sales.csv", sim.sales)
    written["inventory_counts.csv"] = write_csv(PUBLIC / "inventory_counts.csv", counts)
    written["orders.csv"] = write_csv(PUBLIC / "orders.csv", orders)
    written["reservations.csv"] = write_csv(PUBLIC / "reservations.csv", sim.reservations)
    written["shift_reports.csv"] = write_csv(PUBLIC / "shift_reports.csv", shift_reports)
    written["whatsapp_messages.csv"] = write_csv(PUBLIC / "whatsapp_messages.csv", whatsapp)
    written["broadcasts.csv"] = write_csv(PUBLIC / "broadcasts.csv", broadcast_rows)
    if weather_rows:
        written["weather.csv"] = write_csv(PUBLIC / "weather.csv", weather_rows)
    if holiday_rows:
        written["holidays.csv"] = write_csv(PUBLIC / "holidays.csv", holiday_rows)

    # ------------------------------------------------------------ ground truth
    gt_written = {}
    gt_written["true_inventory.csv"] = write_csv(GT / "true_inventory.csv", sim.gt_stock)
    gt_written["count_errors.csv"] = write_csv(GT / "count_errors.csv", sim._count_gt)
    gt_written["incident_events.csv"] = write_csv(GT / "incident_events.csv", sim.gt_events)
    gt_written["shift_report_truth.csv"] = write_csv(GT / "shift_report_truth.csv", shift_truth)
    gt_written["daily_demand.csv"] = write_csv(GT / "daily_demand.csv", sim.daily)
    gt_written["deliveries_received.csv"] = write_csv(
        GT / "deliveries_received.csv", sim.deliveries_received)

    incident_rows = []
    for i in INC.INCIDENTS:
        r = {k: (v.isoformat() if isinstance(v, datetime.date) else v)
             for k, v in i.items()}
        incident_rows.append(r)
    keys = sorted({k for r in incident_rows for k in r})
    for r in incident_rows:
        for k in keys:
            r.setdefault(k, "")
    gt_written["incidents.csv"] = write_csv(
        GT / "incidents.csv", incident_rows, fieldnames=keys)

    # Discrepancy table: for every product, what the books say versus reality at
    # the anchor. This is the answer key for the reconciliation scenarios.
    as_of = C.ANCHOR.date().isoformat()
    prev = (C.ANCHOR.date() - datetime.timedelta(days=1)).isoformat()
    disc = []
    for p in sim.products:
        b = book_stock(sim, p["product_id"], as_of)
        t = truth_at(sim, p["product_id"], prev)
        if b is None or t is None:
            continue
        disc.append({
            "product_id": p["product_id"], "product_name": p["name"],
            "category": p["category"], **b, "physical_stock": t,
            "gap_units": round(b["book"] - t, 3),
            "gap_is_material": abs(b["book"] - t) >= 0.5,
        })
    disc.sort(key=lambda r: -abs(r["gap_units"]))
    gt_written["anchor_discrepancies.csv"] = write_csv(
        GT / "anchor_discrepancies.csv", disc)

    manifest = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "simulation_window": [C.SIM_START.isoformat(), C.SIM_END.isoformat()],
        "anchor": C.ANCHOR.isoformat(),
        "random_seed": C.RANDOM_SEED,
        "weather_loaded": bool(weather_rows),
        "holidays_loaded": bool(holiday_rows),
        "provenance": {
            "simulated": ["products (extended)", "sales", "inventory_counts",
                          "orders", "reservations", "staff_schedule",
                          "shift_reports", "whatsapp_messages"],
            "real": {
                "broadcasts.csv": "Sport5 / livegames.co.il, source_url per row",
                "weather.csv": "Open-Meteo ERA5 archive (keyless)",
                "holidays.csv": "Hebcal REST API (keyless, CC-BY-4.0)",
            },
            "carried_over_from_course_data": [
                "product names, categories, unit costs and prices",
                "shift report phrasing patterns and ambiguity styles",
                "Hebrew WhatsApp slang and message shapes",
            ],
        },
        "public_rows": written,
        "ground_truth_rows": gt_written,
    }
    (ROOT / "data" / "MANIFEST.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    print("\npublic:")
    for k, v in written.items():
        print(f"  {k:26s} {v:>7,}")
    print("ground truth:")
    for k, v in gt_written.items():
        print(f"  {k:26s} {v:>7,}")


if __name__ == "__main__":
    main()
